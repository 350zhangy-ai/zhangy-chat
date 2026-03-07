#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
数据存储与管理模块 - 本地存储、备份、导出
"""

import json
import shutil
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional


class DataManager:
    """数据管理器"""
    
    def __init__(self, data_dir: str = "data", backup_dir: str = "backups"):
        self.data_dir = Path(data_dir)
        self.backup_dir = Path(backup_dir)
        self.data_dir.mkdir(parents=True, exist_ok=True)
        self.backup_dir.mkdir(parents=True, exist_ok=True)
        
        self.config_file = self.data_dir / "config.json"
        self.favorites_file = self.data_dir / "favorites.json"
        self.chat_history_file = self.data_dir / "chat_history.json"
        
        self.config: Dict = {}
        self.favorites: List[Dict] = []
        self.chat_history: List[Dict] = []
        
        self._load_data()
    
    def _load_data(self):
        """加载数据"""
        if self.config_file.exists():
            with open(self.config_file, 'r', encoding='utf-8') as f:
                self.config = json.load(f)
        
        if self.favorites_file.exists():
            with open(self.favorites_file, 'r', encoding='utf-8') as f:
                self.favorites = json.load(f)
        
        if self.chat_history_file.exists():
            with open(self.chat_history_file, 'r', encoding='utf-8') as f:
                self.chat_history = json.load(f)
    
    def _save_data(self):
        """保存数据"""
        with open(self.config_file, 'w', encoding='utf-8') as f:
            json.dump(self.config, f, ensure_ascii=False, indent=2)
        
        with open(self.favorites_file, 'w', encoding='utf-8') as f:
            json.dump(self.favorites, f, ensure_ascii=False, indent=2)
        
        with open(self.chat_history_file, 'w', encoding='utf-8') as f:
            json.dump(self.chat_history, f, ensure_ascii=False, indent=2)
    
    # 配置管理
    def get_config(self, key: str, default=None):
        """获取配置项"""
        return self.config.get(key, default)
    
    def set_config(self, key: str, value):
        """设置配置项"""
        self.config[key] = value
        self._save_data()
    
    def get_all_config(self) -> Dict:
        """获取所有配置"""
        return self.config
    
    # 收藏夹管理
    def add_favorite(self, question: str, answer: str, tags: Optional[List[str]] = None) -> Dict:
        """添加收藏"""
        favorite = {
            "id": datetime.now().strftime("%Y%m%d%H%M%S"),
            "question": question,
            "answer": answer,
            "tags": tags or [],
            "created_at": datetime.now().isoformat()
        }
        self.favorites.append(favorite)
        self._save_data()
        return favorite
    
    def remove_favorite(self, fav_id: str) -> bool:
        """删除收藏"""
        for i, fav in enumerate(self.favorites):
            if fav["id"] == fav_id:
                self.favorites.pop(i)
                self._save_data()
                return True
        return False
    
    def get_favorites(self, tag: Optional[str] = None) -> List[Dict]:
        """获取收藏列表"""
        if tag:
            return [f for f in self.favorites if tag in f.get("tags", [])]
        return self.favorites
    
    def search_favorites(self, keyword: str) -> List[Dict]:
        """搜索收藏"""
        return [f for f in self.favorites 
                if keyword.lower() in f["question"].lower() 
                or keyword.lower() in f["answer"].lower()]
    
    # 聊天记录管理
    def add_chat_record(self, query: str, response: str, 
                        category: str = "general") -> Dict:
        """添加聊天记录"""
        record = {
            "id": datetime.now().strftime("%Y%m%d%H%M%S%f"),
            "query": query,
            "response": response,
            "category": category,
            "timestamp": datetime.now().isoformat()
        }
        self.chat_history.append(record)
        self._save_data()
        return record
    
    def get_chat_history(self, days: int = 7, 
                         category: Optional[str] = None) -> List[Dict]:
        """获取聊天记录"""
        from datetime import timedelta
        cutoff = datetime.now() - timedelta(days=days)
        
        result = self.chat_history
        if category:
            result = [r for r in result if r.get("category") == category]
        
        return [r for r in result 
                if datetime.fromisoformat(r["timestamp"]) >= cutoff]
    
    def clear_chat_history(self, days: Optional[int] = None) -> int:
        """清除聊天记录"""
        if days is None:
            count = len(self.chat_history)
            self.chat_history = []
            self._save_data()
            return count
        
        from datetime import timedelta
        cutoff = datetime.now() - timedelta(days=days)
        original_count = len(self.chat_history)
        self.chat_history = [r for r in self.chat_history 
                            if datetime.fromisoformat(r["timestamp"]) >= cutoff]
        self._save_data()
        return original_count - len(self.chat_history)
    
    # 备份功能
    def create_backup(self, backup_name: Optional[str] = None) -> str:
        """创建备份"""
        if backup_name is None:
            backup_name = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        backup_path = self.backup_dir / backup_name
        backup_path.mkdir(parents=True, exist_ok=True)
        
        # 复制所有数据文件
        for file in self.data_dir.glob("*"):
            if file.is_file():
                shutil.copy2(file, backup_path / file.name)
        
        # 创建备份元数据
        metadata = {
            "backup_name": backup_name,
            "created_at": datetime.now().isoformat(),
            "files": [f.name for f in backup_path.glob("*")]
        }
        with open(backup_path / "metadata.json", 'w', encoding='utf-8') as f:
            json.dump(metadata, f, ensure_ascii=False, indent=2)
        
        return str(backup_path)
    
    def list_backups(self) -> List[Dict]:
        """列出所有备份"""
        backups = []
        for backup_path in self.backup_dir.iterdir():
            if backup_path.is_dir():
                metadata_file = backup_path / "metadata.json"
                if metadata_file.exists():
                    with open(metadata_file, 'r', encoding='utf-8') as f:
                        metadata = json.load(f)
                        backups.append(metadata)
                else:
                    backups.append({
                        "backup_name": backup_path.name,
                        "created_at": "未知",
                        "files": [f.name for f in backup_path.glob("*")]
                    })
        return sorted(backups, key=lambda x: x.get("created_at", ""), reverse=True)
    
    def restore_backup(self, backup_name: str) -> bool:
        """恢复备份"""
        backup_path = self.backup_dir / backup_name
        if not backup_path.exists():
            return False
        
        # 复制备份文件回数据目录
        for file in backup_path.glob("*"):
            if file.name != "metadata.json" and file.is_file():
                shutil.copy2(file, self.data_dir / file.name)
        
        # 重新加载数据
        self._load_data()
        return True
    
    def delete_backup(self, backup_name: str) -> bool:
        """删除备份"""
        backup_path = self.backup_dir / backup_name
        if backup_path.exists() and backup_path.is_dir():
            shutil.rmtree(backup_path)
            return True
        return False
    
    # 导出功能
    def export_to_txt(self, output_path: str, data_type: str = "all") -> str:
        """导出为 TXT 文件"""
        output = Path(output_path)
        lines = []
        
        if data_type in ["tasks", "all"]:
            tasks_file = self.data_dir / "tasks.json"
            if tasks_file.exists():
                with open(tasks_file, 'r', encoding='utf-8') as f:
                    tasks = json.load(f)
                    lines.append("=== 任务列表 ===\n")
                    for task in tasks:
                        status_map = {"pending": "待完成", "completed": "已完成", 
                                     "in_progress": "进行中", "cancelled": "已取消"}
                        lines.append(f"[{status_map.get(task['status'], '未知')}] {task['title']}")
                        if task.get('deadline'):
                            lines.append(f"  截止日期：{task['deadline']}")
                        lines.append("")
        
        if data_type in ["goals", "all"]:
            goals_file = self.data_dir / "goals.json"
            if goals_file.exists():
                with open(goals_file, 'r', encoding='utf-8') as f:
                    goals = json.load(f)
                    lines.append("\n=== 目标列表 ===\n")
                    for goal in goals:
                        lines.append(f"[{goal['status']}] {goal['title']}")
                        if goal.get('milestones'):
                            completed = sum(1 for m in goal['milestones'] if m.get('completed'))
                            lines.append(f"  进度：{completed}/{len(goal['milestones'])} 里程碑")
                        lines.append("")
        
        if data_type in ["habits", "all"]:
            habits_file = self.data_dir / "habits.json"
            if habits_file.exists():
                with open(habits_file, 'r', encoding='utf-8') as f:
                    habits = json.load(f)
                    lines.append("\n=== 习惯打卡 ===\n")
                    for habit in habits:
                        lines.append(f"{habit['title']} - 连续{habit['streak']}天，总计{habit['total_completions']}次")
                    lines.append("")
        
        if data_type in ["favorites", "all"]:
            lines.append("\n=== 收藏列表 ===\n")
            for fav in self.favorites:
                lines.append(f"Q: {fav['question']}")
                lines.append(f"A: {fav['answer'][:100]}..." if len(fav['answer']) > 100 else f"A: {fav['answer']}")
                lines.append("")
        
        with open(output, 'w', encoding='utf-8') as f:
            f.write("\n".join(lines))
        
        return str(output)
    
    def export_to_excel(self, output_path: str, data_type: str = "all") -> str:
        """导出为 Excel 文件"""
        try:
            import openpyxl
            from openpyxl import Workbook
        except ImportError:
            raise ImportError("需要安装 openpyxl: pip install openpyxl")
        
        output = Path(output_path)
        wb = Workbook()
        
        # 删除默认 sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        if data_type in ["tasks", "all"]:
            tasks_file = self.data_dir / "tasks.json"
            if tasks_file.exists():
                with open(tasks_file, 'r', encoding='utf-8') as f:
                    tasks = json.load(f)
                    ws = wb.create_sheet("任务列表")
                    ws.append(["ID", "标题", "描述", "优先级", "状态", "截止日期", "创建时间"])
                    for task in tasks:
                        ws.append([
                            task['id'],
                            task['title'],
                            task.get('description', ''),
                            task.get('priority', 3),
                            task.get('status', 'pending'),
                            task.get('deadline', ''),
                            task.get('created_at', '')
                        ])
        
        if data_type in ["goals", "all"]:
            goals_file = self.data_dir / "goals.json"
            if goals_file.exists():
                with open(goals_file, 'r', encoding='utf-8') as f:
                    goals = json.load(f)
                    ws = wb.create_sheet("目标列表")
                    ws.append(["ID", "标题", "描述", "状态", "进度%", "截止日期", "创建时间"])
                    for goal in goals:
                        completed = sum(1 for m in goal.get('milestones', []) if m.get('completed'))
                        total = len(goal.get('milestones', []))
                        progress = (completed / total * 100) if total > 0 else 0
                        ws.append([
                            goal['id'],
                            goal['title'],
                            goal.get('description', ''),
                            goal.get('status', 'active'),
                            f"{progress:.1f}",
                            goal.get('deadline', ''),
                            goal.get('created_at', '')
                        ])
        
        if data_type in ["habits", "all"]:
            habits_file = self.data_dir / "habits.json"
            if habits_file.exists():
                with open(habits_file, 'r', encoding='utf-8') as f:
                    habits = json.load(f)
                    ws = wb.create_sheet("习惯打卡")
                    ws.append(["ID", "名称", "频率", "连续天数", "总完成次数", "创建时间"])
                    for habit in habits:
                        ws.append([
                            habit['id'],
                            habit['title'],
                            habit.get('frequency', 'daily'),
                            habit.get('streak', 0),
                            habit.get('total_completions', 0),
                            habit.get('created_at', '')
                        ])
        
        if data_type in ["favorites", "all"]:
            ws = wb.create_sheet("收藏列表")
            ws.append(["ID", "问题", "答案", "标签", "创建时间"])
            for fav in self.favorites:
                ws.append([
                    fav['id'],
                    fav['question'],
                    fav['answer'],
                    ", ".join(fav.get('tags', [])),
                    fav.get('created_at', '')
                ])
        
        wb.save(output)
        return str(output)
    
    def export_to_csv(self, output_path: str, data_type: str = "tasks") -> str:
        """导出为 CSV 文件"""
        output = Path(output_path)
        
        if data_type == "tasks":
            tasks_file = self.data_dir / "tasks.json"
            if not tasks_file.exists():
                raise FileNotFoundError("任务数据不存在")
            
            with open(tasks_file, 'r', encoding='utf-8') as f:
                tasks = json.load(f)
            
            with open(output, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=['id', 'title', 'description', 
                                                       'priority', 'status', 'deadline', 
                                                       'created_at', 'completed_at'])
                writer.writeheader()
                for task in tasks:
                    writer.writerow({k: task.get(k, '') for k in writer.fieldnames})
        
        elif data_type == "goals":
            goals_file = self.data_dir / "goals.json"
            if not goals_file.exists():
                raise FileNotFoundError("目标数据不存在")
            
            with open(goals_file, 'r', encoding='utf-8') as f:
                goals = json.load(f)
            
            with open(output, 'w', encoding='utf-8-sig', newline='') as f:
                writer = csv.DictWriter(f, fieldnames=['id', 'title', 'description',
                                                       'status', 'deadline', 'created_at'])
                writer.writeheader()
                for goal in goals:
                    writer.writerow({k: goal.get(k, '') for k in writer.fieldnames})
        
        return str(output)
